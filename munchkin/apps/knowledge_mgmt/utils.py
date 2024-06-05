import openpyxl


def get_index_name(knowledge_base_folder_id):
    return f"knowledge_base_{knowledge_base_folder_id}"


def excel_to_dict_all_sheets(excel_file_path, chunk_size=1000):
    # 打开Excel文件
    wb = openpyxl.load_workbook(excel_file_path, read_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # 获取表头
        header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]

        # 存储当前工作表结果的列表
        chunk = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if set(row) == {None}:
                continue
            chunk.append(dict(zip(header, row)))

            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []
        else:
            yield chunk

    wb.close()
